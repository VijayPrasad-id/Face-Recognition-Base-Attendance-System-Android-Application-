# main.py
"""
Face Attendance Kivy app (Buildozer-friendly).

Notes:
- Replaces bcrypt with PBKDF2 (hashlib) to avoid C-extension builds.
- Uses Kivy Camera on Android; uses cv2.VideoCapture on desktop.
- Guards OpenCV usage with OPENCV_AVAILABLE flag so packaging won't fail
  if opencv-python is not present during build.
- openpyxl usage is also guarded.
"""

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.textinput import TextInput
from kivy.uix.image import Image
from kivy.clock import Clock
from kivy.graphics.texture import Texture
from kivy.utils import platform
from kivy.uix.camera import Camera

import os
import csv
import datetime
import hashlib
import binascii
import numpy as np

# Try OpenCV; if unavailable, set OPENCV_AVAILABLE = False
try:
    import cv2
    OPENCV_AVAILABLE = True
except Exception:
    cv2 = None
    OPENCV_AVAILABLE = False

# Try openpyxl; guard Excel export if not available
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    openpyxl = None
    get_column_letter = None
    OPENPYXL_AVAILABLE = False

# ----------------- PATHS & PERMISSIONS -----------------

if platform == "android":
    try:
        from android.permissions import request_permissions, Permission  # type: ignore
        request_permissions(
            [
                Permission.CAMERA,
                Permission.WRITE_EXTERNAL_STORAGE,
                Permission.READ_EXTERNAL_STORAGE,
            ]
        )
    except Exception:
        # If permissions module not available during build, ignore
        pass

    DATA_DIR = App.get_running_app().user_data_dir if App.get_running_app() else "data"
else:
    DATA_DIR =r"C:\Users\vijay\Downloads\sem 6\Project\NACL MANU\data"

FACES_DIR = os.path.join(DATA_DIR, "faces")
PEOPLE_CSV = os.path.join(DATA_DIR, "people.csv")
ATTENDANCE_CSV = os.path.join(DATA_DIR, "attendance.csv")
TRAINER_FILE = os.path.join(DATA_DIR, "trainer.yml")
ROOT_FILE = os.path.join(DATA_DIR, "root_user.txt")

os.makedirs(FACES_DIR, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)

# Initialize CSV files if missing
if not os.path.exists(PEOPLE_CSV):
    with open(PEOPLE_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "name"])

if not os.path.exists(ATTENDANCE_CSV):
    with open(ATTENDANCE_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "name", "date", "login_time", "logout_time"])

# Face cascade only if OpenCV present
if OPENCV_AVAILABLE:
    face_cascade = cv2.CascadeClassifier(
        cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
    )
else:
    face_cascade = None

# ----------------- PBKDF2 helpers (replace bcrypt) -----------------


def hash_password(password, salt=None, iterations=200_000):
    """
    Return (salt_hex, dk_hex) for the given password.
    Uses PBKDF2-HMAC-SHA256; pure Python (no C ext).
    """
    if salt is None:
        salt = os.urandom(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    return binascii.hexlify(salt).decode(), binascii.hexlify(dk).decode()


def verify_password(password, salt_hex, dk_hex, iterations=200_000):
    salt = binascii.unhexlify(salt_hex)
    dk_check = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    return binascii.hexlify(dk_check).decode() == dk_hex


# ----------------- CSV / File helpers -----------------


def load_people():
    people = {}
    if os.path.exists(PEOPLE_CSV):
        with open(PEOPLE_CSV, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    people[int(row["id"])] = row["name"]
                except Exception:
                    continue
    return people


def save_person(person_id, name):
    people = load_people()
    people[person_id] = name
    with open(PEOPLE_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "name"])
        for pid, pname in sorted(people.items()):
            writer.writerow([pid, pname])


def load_attendance():
    if os.path.exists(ATTENDANCE_CSV):
        with open(ATTENDANCE_CSV, newline="", encoding="utf-8") as f:
            return list(csv.DictReader(f))
    return []


def save_attendance(records):
    with open(ATTENDANCE_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["id", "name", "date", "login_time", "logout_time"],
        )
        writer.writeheader()
        writer.writerows(records)


def show_popup(title, message):
    popup = Popup(
        title=title,
        content=Label(text=message),
        size_hint=(None, None),
        size=(400, 200),
    )
    popup.open()


# ---------- Root user helpers (PBKDF2) ----------


def check_root_credentials(username, password):
    if not os.path.exists(ROOT_FILE):
        return False
    try:
        with open(ROOT_FILE, "r", encoding="utf-8") as f:
            lines = f.read().splitlines()
        if len(lines) < 2:
            return False
        stored_user = lines[0].strip()
        salt_dk = lines[1].strip()
        salt_hex, dk_hex = salt_dk.split(":")
        if username != stored_user:
            return False
        return verify_password(password, salt_hex, dk_hex)
    except Exception:
        return False


def create_root_user_if_needed():
    """
    If ROOT_FILE missing, prompt user at startup to create credentials.
    Uses PBKDF2 to store salted hash in plain text file.
    """
    if not os.path.exists(ROOT_FILE):
        from kivy.uix.gridlayout import GridLayout

        layout = GridLayout(cols=1, spacing=10, padding=20)
        user_input = TextInput(hint_text="Set root username")
        pass_input = TextInput(hint_text="Set root password", password=True)
        btn = Button(text="Save root user")

        layout.add_widget(user_input)
        layout.add_widget(pass_input)
        layout.add_widget(btn)

        popup = Popup(title="Create Root User", content=layout, size_hint=(0.8, 0.6))

        def save_root(instance):
            u = user_input.text.strip()
            p = pass_input.text.strip()
            if not u or not p:
                show_popup("Error", "Username and password required")
                return
            salt_hex, dk_hex = hash_password(p)
            with open(ROOT_FILE, "w", encoding="utf-8") as f:
                f.write(u + "\n")
                f.write(salt_hex + ":" + dk_hex + "\n")
            popup.dismiss()
            app = App.get_running_app()
            if app:
                app.show_login_screen()

        btn.bind(on_press=save_root)
        popup.open()
    else:
        app = App.get_running_app()
        if app:
            app.show_login_screen()


# ---------- Excel helpers (guarded) ----------


def get_month_excel_path(year, month):
    fname = f"attendance_{year}_{month:02d}.xlsx"
    return os.path.join(DATA_DIR, fname)


def write_month_to_excel(year, month):
    if not OPENPYXL_AVAILABLE:
        show_popup("Error", "openpyxl not available")
        return
    records = load_attendance()
    filtered = []
    for r in records:
        try:
            d = datetime.datetime.strptime(r["date"], "%Y-%m-%d").date()
        except Exception:
            continue
        if d.year == year and d.month == month:
            filtered.append(r)

    if not filtered:
        show_popup("Info", "No attendance records for this month.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    headers = ["id", "name", "date", "login_time", "logout_time"]
    ws.append(headers)
    for r in filtered:
        ws.append([r["id"], r["name"], r["date"], r["login_time"], r["logout_time"]])

    for col_idx, col_name in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for row in ws.iter_rows(
            min_row=2, min_col=col_idx, max_col=col_idx, values_only=True
        ):
            cell_val = row[0]
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = max_len + 2

    path = get_month_excel_path(year, month)
    wb.save(path)
    show_popup("Success", f"Excel saved:\n{path}")


# ---------- User management helpers ----------


def delete_person_from_csv(person_id):
    people = load_people()
    if person_id in people:
        del people[person_id]
        with open(PEOPLE_CSV, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["id", "name"])
            for pid, pname in sorted(people.items()):
                writer.writerow([pid, pname])
        return True
    return False


def delete_person_faces(person_id):
    if not os.path.exists(FACES_DIR):
        return
    for fname in os.listdir(FACES_DIR):
        if fname.startswith(f"{person_id}_"):
            try:
                os.remove(os.path.join(FACES_DIR, fname))
            except OSError:
                pass


# ---------- Duplicate face helpers (guarded) ----------


def predict_existing_id_for_face(face_img):
    """
    Given a 200x200 grayscale face image, try to predict existing ID.
    Returns (label, confidence) or (None, None) if model not available.
    """
    if not OPENCV_AVAILABLE or not os.path.exists(TRAINER_FILE):
        return None, None
    try:
        recognizer = cv2.face.LBPHFaceRecognizer_create()
        recognizer.read(TRAINER_FILE)
        label, confidence = recognizer.predict(face_img)
        return label, confidence
    except Exception:
        return None, None


# ----------------- Cross-platform camera widget -----------------


class CrossPlatformCamera(Image):
    def __init__(self, mode="recognize", person_id=None, name=None, **kwargs):
        super(CrossPlatformCamera, self).__init__(**kwargs)
        self.mode = mode
        self.person_id = person_id
        self.name = name
        self.sample_count = 0
        self.recognized_id = None
        self.start_time = datetime.datetime.now()
        self.on_android = platform == "android"
        self.recognizer = None

        if OPENCV_AVAILABLE and mode == "recognize" and os.path.exists(TRAINER_FILE):
            try:
                self.recognizer = cv2.face.LBPHFaceRecognizer_create()
                self.recognizer.read(TRAINER_FILE)
            except Exception:
                self.recognizer = None

        if self.on_android:
            self.kivy_cam = Camera(index=0, resolution=(640, 480), play=True)
            Clock.schedule_interval(self.update_from_kivy_cam, 1.0 / 20.0)
        else:
            if OPENCV_AVAILABLE:
                self.cap = cv2.VideoCapture(0)
                self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
                self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
            else:
                self.cap = None
            Clock.schedule_interval(self.update_from_cv, 1.0 / 30.0)

    def update_from_kivy_cam(self, dt):
        tex = self.kivy_cam.texture
        if not tex:
            return
        buf = tex.pixels
        w, h = tex.size
        arr = np.frombuffer(buf, np.uint8)
        try:
            arr = arr.reshape(int(h), int(w), 4)
        except Exception:
            return
        frame = arr[:, :, :3]
        frame = frame[:, :, ::-1]
        self.process_frame(frame)

    def update_from_cv(self, dt):
        if not OPENCV_AVAILABLE or self.cap is None:
            return
        ret, frame = self.cap.read()
        if not ret:
            return
        self.process_frame(frame)

    def process_frame(self, frame):
        frame_small = frame.copy()
        if OPENCV_AVAILABLE and frame_small.shape[1] > 320:
            frame_small = cv2.resize(frame_small, (320, 240))
        if OPENCV_AVAILABLE and face_cascade is not None:
            gray_small = cv2.cvtColor(frame_small, cv2.COLOR_BGR2GRAY)
            faces_small = face_cascade.detectMultiScale(
                gray_small, scaleFactor=1.2, minNeighbors=5, minSize=(50, 50)
            )
        else:
            faces_small = []

        scale_x = frame.shape[1] / max(1, frame_small.shape[1])
        scale_y = frame.shape[0] / max(1, frame_small.shape[0])

        if self.mode == "register" and OPENCV_AVAILABLE:
            gray_full = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            for (x_s, y_s, w_s, h_s) in faces_small:
                x = int(x_s * scale_x)
                y = int(y_s * scale_y)
                w = int(w_s * scale_x)
                h = int(h_s * scale_y)
                self.sample_count += 1
                face_img = cv2.resize(gray_full[y : y + h, x : x + w], (200, 200))
                cv2.imwrite(
                    os.path.join(FACES_DIR, f"{self.person_id}_{self.sample_count}.jpg"),
                    face_img,
                )
                cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 0), 2)
                cv2.putText(
                    frame,
                    f"Sample {self.sample_count}/30",
                    (x, y - 10),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.7,
                    (0, 255, 0),
                    2,
                )
            if self.sample_count >= 30:
                self.stop()
                save_person(self.person_id, self.name)
                show_popup("Success", "Samples collected. Please train the model.")

        elif self.mode == "recognize" and OPENCV_AVAILABLE:
            gray_full = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            for (x_s, y_s, w_s, h_s) in faces_small:
                x = int(x_s * scale_x)
                y = int(y_s * scale_y)
                w = int(w_s * scale_x)
                h = int(h_s * scale_y)
                face_img = cv2.resize(gray_full[y : y + h, x : x + w], (200, 200))
                if self.recognizer is not None:
                    try:
                        label, confidence = self.recognizer.predict(face_img)
                        if confidence < 70:
                            self.recognized_id = label
                            self.stop()
                            App.get_running_app().process_attendance(self.recognized_id)
                            return
                    except Exception:
                        pass
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
            if (datetime.datetime.now() - self.start_time).seconds > 10:
                self.stop()
                show_popup("Error", "Face not recognized within 10 seconds.")

        # Convert frame to texture safely
        try:
            buf1 = cv2.flip(frame, 0) if OPENCV_AVAILABLE else frame
            buf = buf1.tobytes()
            texture = Texture.create(size=(frame.shape[1], frame.shape[0]), colorfmt="bgr")
            texture.blit_buffer(buf, colorfmt="bgr", bufferfmt="ubyte")
            self.texture = texture
        except Exception:
            # If conversion fails, ignore to avoid crashes on devices without cv2
            pass

    def stop(self):
        Clock.unschedule(
            self.update_from_kivy_cam if self.on_android else self.update_from_cv
        )
        if not self.on_android and OPENCV_AVAILABLE and hasattr(self, "cap"):
            try:
                self.cap.release()
            except Exception:
                pass
        if self.on_android and hasattr(self, "kivy_cam"):
            try:
                self.kivy_cam.play = False
            except Exception:
                pass


# ----------------- MAIN APP UI -----------------


class AttendanceApp(App):
    def build(self):
        self.action = None
        self.root_layout = BoxLayout(orientation="vertical")
        Clock.schedule_once(lambda dt: create_root_user_if_needed(), 0)
        return self.root_layout

    # ----- Login / main menu -----

    def show_login_screen(self):
        self.root_layout.clear_widgets()
        layout = BoxLayout(orientation="vertical", padding=20, spacing=10)
        title = Label(text="Root Login", font_size=24, size_hint=(1, 0.2))
        self.login_user = TextInput(hint_text="Username")
        self.login_pass = TextInput(hint_text="Password", password=True)
        btn = Button(text="Login")

        layout.add_widget(title)
        layout.add_widget(self.login_user)
        layout.add_widget(self.login_pass)
        layout.add_widget(btn)

        btn.bind(on_press=self.perform_login)
        self.root_layout.add_widget(layout)

    def perform_login(self, instance):
        u = self.login_user.text.strip()
        p = self.login_pass.text.strip()
        if check_root_credentials(u, p):
            self.show_main_menu()
        else:
            show_popup("Error", "Invalid username or password")

    def show_main_menu(self):
        self.root_layout.clear_widgets()
        layout = BoxLayout(orientation="vertical", padding=20, spacing=20)

        avatar_path = os.path.join(DATA_DIR, "avta.png")
        if os.path.exists(avatar_path):
            avatar = Image(
                source=avatar_path, 
                size_hint=(1, 0.35),
            )
            layout.add_widget(avatar)
            title_size_hint = (1, 0.1)
        else:
            title_size_hint = (1, 0.15)

        title = Label(
            text="[b][color=ffffff]Welcome to Face Attendance System[/color][/b]",
            font_size=22,
            size_hint=(1, 0.1),
            markup=True,
        )
        
        layout.add_widget(title)
        

        row1 = GridLayout(cols=3, spacing=15, size_hint=(1, 0.2))
        new_img = os.path.join(DATA_DIR, "new.png")
        btn_register = Button(
            text="New\nRegistration",
            font_size=14,
            bold=True,
        )
        if os.path.exists(new_img):
            btn_register.background_normal = new_img
            btn_register.background_down = new_img
            btn_register.color = (0, 0, 0, 1)
        else:
            btn_register.background_color = (0.2, 0.6, 1, 1)  # Blue
            btn_register.color = (1, 1, 1, 1)  # White text
        btn_register.bind(on_press=self.ui_register)


        log_img = os.path.join(DATA_DIR, "log.png")
        btn_login = Button(
            text="Login",
            font_size=14,
            bold=True,
        )
        if os.path.exists(log_img):
            btn_login.background_normal = log_img
            btn_login.background_down = log_img
            btn_login.color = (0, 0, 0, 1)
        else:
            btn_login.background_color = (0.2, 0.8, 0.2, 1)  # Green
            btn_login.color = (1, 1, 1, 1)
        btn_login.bind(on_press=lambda x: self.start_recognition("login"))


        logout_img = os.path.join(DATA_DIR, "logout.png")
        btn_logout = Button(
            text="Logout",
            font_size=14,
            bold=True,
        )
        if os.path.exists(logout_img):
            btn_logout.background_normal = logout_img
            btn_logout.background_down = logout_img
            btn_logout.color = (0, 0, 0, 1)
        else:
            btn_logout.background_color = (1, 0.4, 0.4, 1)  # Red
            btn_logout.color = (1, 1, 1, 1)
        btn_logout.bind(on_press=lambda x: self.start_recognition("logout"))

        row1.add_widget(btn_register)
        row1.add_widget(btn_login)
        row1.add_widget(btn_logout)
        layout.add_widget(row1)

        row2 = GridLayout(cols=3, spacing=15, size_hint=(1, 0.2))
        ma_img = os.path.join(DATA_DIR, "ma.png")
        btn_manual = Button(
            text="Manual\nAttendance",
            font_size=14,
            bold=True,
        )
        if os.path.exists(ma_img):
            btn_manual.background_normal = ma_img
            btn_manual.background_down = ma_img
            btn_manual.color = (0, 0, 0, 1)
        else:
            btn_manual.background_color = (1, 0.7, 0.2, 1)  # Orange
            btn_manual.color = (1, 1, 1, 1)
        btn_manual.bind(on_press=self.manual_attendance_ui)


        train_img = os.path.join(DATA_DIR, "train.png")
        btn_train = Button(
            text="Train\nModel",
            font_size=14,
            bold=True,
        )
        if os.path.exists(train_img):
            btn_train.background_normal = train_img
            btn_train.background_down = train_img
            btn_train.color = (0, 0, 0, 1)
        else:
            btn_train.background_color = (0.6, 0.2, 0.8, 1)  # Purple
            btn_train.color = (1, 1, 1, 1)
        btn_train.bind(on_press=self.train_recognizer)


        va_img = os.path.join(DATA_DIR, "va.png")
        btn_view = Button(
            text="View/Save\nAttendance",
            font_size=14,
            bold=True,
        )
        if os.path.exists(va_img):
            btn_view.background_normal = va_img
            btn_view.background_down = va_img
            btn_view.color = (0, 0, 0, 1)
        else:
            btn_view.background_color = (0.2, 0.6, 0.8, 1)  # Cyan
            btn_view.color = (1, 1, 1, 1)
        btn_view.bind(on_press=self.view_attendance_ui)
        row2.add_widget(btn_manual)
        row2.add_widget(btn_train)
        row2.add_widget(btn_view)
        layout.add_widget(row2)

        row3 = GridLayout(cols=3, spacing=15, size_hint=(1, 0.2))
        manage_img = os.path.join(DATA_DIR, "manage.png")
        btn_manage_users = Button(
            text="Manage\nUsers",
            font_size=14,
            bold=True,
        )
        if os.path.exists(manage_img):
            btn_manage_users.background_normal = manage_img
            btn_manage_users.background_down = manage_img
            btn_manage_users.color = (0, 0, 0, 1)
        else:
            btn_manage_users.background_color = (0.8, 0.4, 0.2, 1)  # Brown
            btn_manage_users.color = (1, 1, 1, 1)
        btn_manage_users.bind(on_press=self.manage_users_ui)
        row3.add_widget(btn_manage_users)
        layout.add_widget(row3)

        self.root_layout.add_widget(layout)

    # ----- Registration -----

    def ui_register(self, instance):
        content = BoxLayout(orientation="vertical", spacing=10)
        self.id_input = TextInput(hint_text="Enter Numeric ID")
        self.name_input = TextInput(hint_text="Enter Full Name")
        submit_btn = Button(text="Start Camera")

        content.add_widget(self.id_input)
        content.add_widget(self.name_input)
        content.add_widget(submit_btn)

        popup = Popup(title="Register New Person", content=content, size_hint=(0.8, 0.5))
        submit_btn.bind(on_press=lambda x: self.start_capture(popup))
        popup.open()

    def start_capture(self, popup):
        try:
            p_id = int(self.id_input.text)
            name = self.name_input.text.strip()
            if not name:
                show_popup("Error", "Name cannot be empty.")
                return

            people = load_people()
            if p_id in people:
                show_popup("Error", f"ID {p_id} already exists for {people[p_id]}.")
                return

            for existing_id, existing_name in people.items():
                if existing_name.lower() == name.lower():
                    show_popup(
                        "Error",
                        f"Name '{name}' is already used with ID {existing_id}.",
                    )
                    return

            if os.path.exists(TRAINER_FILE) and OPENCV_AVAILABLE:
                duplicate_id = self.check_duplicate_face_before_register()
                if duplicate_id is not None:
                    existing_name = people.get(duplicate_id, "Unknown")
                    show_popup(
                        "Duplicate face",
                        (
                            f"This face looks like existing user ID {duplicate_id} "
                            f"({existing_name}).\nIf this is the same person, do "
                            "not create a new ID."
                        ),
                    )
                    return

            popup.dismiss()
            cam_popup = Popup(title="Capturing 30 Samples...", size_hint=(0.9, 0.9))
            cam_widget = CrossPlatformCamera(mode="register", person_id=p_id, name=name)
            cam_popup.content = cam_widget
            cam_popup.bind(on_dismiss=lambda *args: cam_widget.stop())
            cam_popup.open()
        except ValueError:
            show_popup("Error", "ID must be a number")

    def check_duplicate_face_before_register(self, samples_to_check=10, threshold=60.0):
        if not OPENCV_AVAILABLE:
            return None
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            return None

        match_counts = {}
        frames_captured = 0

        while frames_captured < samples_to_check:
            ret, frame = cap.read()
            if not ret:
                break
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(
                gray, scaleFactor=1.2, minNeighbors=5, minSize=(100, 100)
            )
            if len(faces) == 0:
                continue

            x, y, w, h = faces[0]
            face_img = cv2.resize(gray[y : y + h, x : x + w], (200, 200))
            label, confidence = predict_existing_id_for_face(face_img)
            if label is not None and confidence < threshold:
                match_counts[label] = match_counts.get(label, 0) + 1
                frames_captured += 1

        cap.release()

        if not match_counts:
            return None

        best_id = max(match_counts, key=match_counts.get)
        best_count = match_counts[best_id]

        if best_count >= samples_to_check // 2:
            return best_id
        return None

    # ----- Training -----

    def train_recognizer(self, instance):
        if not OPENCV_AVAILABLE:
            show_popup("Error", "OpenCV not available")
            return

        faces, labels = [], []
        for filename in os.listdir(FACES_DIR):
            if filename.lower().endswith((".jpg", ".png")):
                path = os.path.join(FACES_DIR, filename)
                img = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
                if img is not None:
                    faces.append(img)
                    labels.append(int(filename.split("_")[0]))

        if faces:
            recognizer = cv2.face.LBPHFaceRecognizer_create()
            recognizer.train(faces, np.array(labels))
            os.makedirs(DATA_DIR, exist_ok=True)
            recognizer.save(TRAINER_FILE)
            show_popup("Success", "Model trained successfully!")
        else:
            show_popup("Error", "No face samples found to train.")

    # ----- Face recognition login/logout -----

    def start_recognition(self, action):
        if OPENCV_AVAILABLE and not os.path.exists(TRAINER_FILE):
            show_popup("Error", "Please register and train the model first.")
            return
        self.action = action
        cam_popup = Popup(title=f"Face Recognition ({action})", size_hint=(0.9, 0.9))
        cam_widget = CrossPlatformCamera(mode="recognize")
        cam_popup.content = cam_widget
        cam_popup.bind(on_dismiss=lambda *args: cam_widget.stop())
        cam_popup.open()

    def process_attendance(self, person_id):
        people = load_people()
        name = people.get(person_id)
        if not name:
            show_popup("Error", "ID not found in database.")
            return

        today = datetime.date.today().isoformat()
        records = load_attendance()

        if self.action == "login":
            for row in records:
                if row["id"] == str(person_id) and row["date"] == today:
                    show_popup("Info", f"{name} is already logged in today.")
                    return
            records.append(
                {
                    "id": str(person_id),
                    "name": name,
                    "date": today,
                    "login_time": datetime.datetime.now().strftime("%H:%M:%S"),
                    "logout_time": "",
                }
            )
            save_attendance(records)
            show_popup("Success", f"Login recorded for {name}")

        elif self.action == "logout":
            found = False
            for row in records:
                if (
                    row["id"] == str(person_id)
                    and row["date"] == today
                    and row["logout_time"] == ""
                ):
                    row["logout_time"] = datetime.datetime.now().strftime("%H:%M:%S")
                    found = True
                    break
            if found:
                save_attendance(records)
                show_popup("Success", f"Logout recorded for {name}")
            else:
                show_popup("Info", f"No active login found for {name} today.")

    # ----- Manual attendance (root only) -----

    def manual_attendance_ui(self, instance):
        people = load_people()
        if not people:
            show_popup("Info", "No people registered.")
            return

        content = BoxLayout(orientation="vertical", spacing=10, padding=10)
        self.manual_id_input = TextInput(hint_text="Person ID (numeric)")
        self.manual_date_input = TextInput(
            hint_text="Date YYYY-MM-DD (empty = today)"
        )
        self.manual_type_input = TextInput(hint_text="Type: login or logout")
        btn = Button(text="Save manual attendance")

        content.add_widget(self.manual_id_input)
        content.add_widget(self.manual_date_input)
        content.add_widget(self.manual_type_input)
        content.add_widget(btn)

        popup = Popup(
            title="Manual Attendance (Root only)", content=content, size_hint=(0.8, 0.6)
        )

        def do_manual(instance_btn):
            try:
                pid = int(self.manual_id_input.text)
            except ValueError:
                show_popup("Error", "ID must be numeric")
                return
            date_text = self.manual_date_input.text.strip()
            if date_text:
                try:
                    dt = datetime.datetime.strptime(date_text, "%Y-%m-%d").date()
                except ValueError:
                    show_popup("Error", "Date format must be YYYY-MM-DD")
                    return
            else:
                dt = datetime.date.today()
            kind = self.manual_type_input.text.strip().lower()
            if kind not in ("login", "logout"):
                show_popup("Error", "Type must be 'login' or 'logout'")
                return

            old_action = self.action
            self.action = kind
            self.process_attendance_manual(pid, dt)
            self.action = old_action

        btn.bind(on_press=do_manual)
        popup.open()

    def process_attendance_manual(self, person_id, date_obj):
        people = load_people()
        name = people.get(person_id)
        if not name:
            show_popup("Error", "ID not found in database.")
            return

        date_str = date_obj.isoformat()
        records = load_attendance()

        if self.action == "login":
            for row in records:
                if row["id"] == str(person_id) and row["date"] == date_str:
                    show_popup("Info", f"{name} is already logged in on {date_str}.")
                    return
            records.append(
                {
                    "id": str(person_id),
                    "name": name,
                    "date": date_str,
                    "login_time": datetime.datetime.now().strftime("%H:%M:%S"),
                    "logout_time": "",
                }
            )
            save_attendance(records)
            show_popup("Success", f"Manual login recorded for {name} on {date_str}")
        elif self.action == "logout":
            found = False
            for row in records:
                if (
                    row["id"] == str(person_id)
                    and row["date"] == date_str
                    and row["logout_time"] == ""
                ):
                    row["logout_time"] = datetime.datetime.now().strftime("%H:%M:%S")
                    found = True
                    break
            if found:
                save_attendance(records)
                show_popup(
                    "Success", f"Manual logout recorded for {name} on {date_str}"
                )
            else:
                show_popup("Info", f"No active login found for {name} on {date_str}.")

    # ----- View / save attendance as Excel -----

    def view_attendance_ui(self, instance):
        content = BoxLayout(orientation="vertical", spacing=10, padding=10)
        self.view_year = TextInput(hint_text="Year (e.g. 2026)")
        self.view_month = TextInput(hint_text="Month (1-12)")
        btn = Button(text="Generate & Save Excel")

        content.add_widget(self.view_year)
        content.add_widget(self.view_month)
        content.add_widget(btn)

        popup = Popup(title="View / Save Attendance (Excel)", content=content, size_hint=(0.8, 0.6))

        def gen(instance_btn):
            try:
                y = int(self.view_year.text)
                m = int(self.view_month.text)
                if not (1 <= m <= 12):
                    raise ValueError
            except ValueError:
                show_popup("Error", "Enter valid year and month")
                return
            write_month_to_excel(y, m)

        btn.bind(on_press=gen)
        popup.open()

    # ----- Manage users (delete) -----

    def manage_users_ui(self, instance):
        people = load_people()
        if not people:
            show_popup("Info", "No users to manage.")
            return

        content = BoxLayout(orientation="vertical", spacing=10, padding=10)
        info_label = Label(text="Enter ID to delete user (this removes face data).")
        self.delete_id_input = TextInput(hint_text="User ID (numeric)")
        btn_delete = Button(text="Delete user")

        content.add_widget(info_label)
        content.add_widget(self.delete_id_input)
        content.add_widget(btn_delete)

        popup = Popup(title="Manage Users (Root only)", content=content, size_hint=(0.8, 0.6))

        def do_delete(instance_btn):
            try:
                pid = int(self.delete_id_input.text)
            except ValueError:
                show_popup("Error", "ID must be numeric.")
                return
            people_local = load_people()
            if pid not in people_local:
                show_popup("Error", "ID not found.")
                return
            name = people_local[pid]
            delete_person_faces(pid)
            delete_person_from_csv(pid)
            show_popup("Success", f"User {name} (ID {pid}) deleted.")
            popup.dismiss()

        btn_delete.bind(on_press=do_delete)
        popup.open()


if __name__ == "__main__":
    AttendanceApp().run()